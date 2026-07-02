"""Refresh model-service scenario fixtures from the source Excel workbooks.

The generated JSON is intentionally committed so tests do not depend on the
operator's KakaoTalk download directory. Run this script manually when the
source workbooks change.
"""

from __future__ import annotations

import argparse
import json
import re
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover - operator setup guard
    raise SystemExit(
        "openpyxl is required to refresh scenario fixtures. "
        "Use the bundled Codex Python runtime or install test tooling first."
    ) from exc


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUTPUT = ROOT / "services/model/tests/fixtures/scenario_matrix.json"
DEFAULT_REPORT = ROOT / "docs/scenario-readiness/scenario_fixture_report.md"

SCENARIO_COLUMNS = [
    "block_id",
    "case_id",
    "level_id",
    "person_count",
    "pick_position",
    "hand_or_person",
    "return_yn",
    "return_position",
    "product_mix",
    "single_pick_count",
    "total_pick_count",
    "speed",
    "door_open",
    "return_timing",
    "return_speed",
    "expected_basket",
    "check_item",
    "notes",
    "test_result_note",
    "execution_status",
    "assignee",
    "executed_at",
    "issue_or_note",
]

CHECKLIST_COLUMNS = [
    "item_no",
    "category",
    "payment_method",
    "test_content",
    "elapsed_time",
    "amount",
    "final_result",
    "notes",
]

PRODUCT_IDS = {"A": 101, "B": 102, "C": 103}


def _value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        return stripped if stripped else None
    return value


def _find_sheet_by_shape(workbook: Any, *, min_rows: int, min_cols: int) -> Any:
    candidates = [
        sheet
        for sheet in workbook.worksheets
        if sheet.max_row >= min_rows and sheet.max_column >= min_cols
    ]
    if not candidates:
        raise ValueError(f"No worksheet found with shape >= {min_rows}x{min_cols}")
    return max(candidates, key=lambda sheet: (sheet.max_row, sheet.max_column))


def _parse_expected_counts(expected_basket: str | None) -> dict[str, int]:
    if not expected_basket:
        return {}
    if "없음" in expected_basket:
        return {}
    counts: Counter[str] = Counter()
    for product_key, count_text in re.findall(r"([ABC])\s*(\d+)\s*개", expected_basket):
        counts[product_key] += int(count_text)
    return dict(sorted(counts.items()))


def _case_coverage(case: dict[str, Any]) -> str:
    expected = case.get("expected_counts") or {}
    if expected:
        return "model_contract"
    if "없음" in str(case.get("expected_basket") or ""):
        return "model_contract_empty"
    return "external_or_manual_review"


def load_scenario_cases(path: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    summary_sheet = workbook.worksheets[0]
    case_sheet = _find_sheet_by_shape(workbook, min_rows=900, min_cols=20)

    block_rows: list[dict[str, Any]] = []
    for row in summary_sheet.iter_rows(min_row=2, values_only=True):
        block_id = _value(row[0] if len(row) > 0 else None)
        if not block_id:
            continue
        block_rows.append(
            {
                "block_id": block_id,
                "scenario_name": _value(row[1] if len(row) > 1 else None),
                "guide": _value(row[2] if len(row) > 2 else None),
                "condition": _value(row[3] if len(row) > 3 else None),
            }
        )

    cases: list[dict[str, Any]] = []
    current_block_id: str | None = None
    for row in case_sheet.iter_rows(min_row=4, values_only=True):
        values = [_value(cell) for cell in row[: len(SCENARIO_COLUMNS)]]
        raw_block = values[0]
        case_id = values[1]
        if isinstance(raw_block, str) and re.fullmatch(r"S\d{2}", raw_block):
            current_block_id = raw_block
        if not case_id or case_id == "Case ID":
            continue
        if not current_block_id:
            raise ValueError(f"Case {case_id} is missing a block id")
        values[0] = current_block_id
        case = dict(zip(SCENARIO_COLUMNS, values))
        expected_counts = _parse_expected_counts(case.get("expected_basket"))
        case["expected_counts"] = expected_counts
        case["expected_product_ids"] = {
            product_key: PRODUCT_IDS[product_key] for product_key in expected_counts
        }
        case["allows_exception"] = "예외" in str(case.get("expected_basket") or "")
        case["coverage"] = _case_coverage(case)
        cases.append(case)

    return block_rows, cases


def load_checklist_rows(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = _find_sheet_by_shape(workbook, min_rows=100, min_cols=8)
    rows: list[dict[str, Any]] = []
    category: Any = None
    payment_method: Any = None
    for row in sheet.iter_rows(min_row=2, values_only=True):
        values = [_value(cell) for cell in row[: len(CHECKLIST_COLUMNS)]]
        if values[0] is None:
            continue
        if values[1] is not None:
            category = values[1]
        if values[2] is not None:
            payment_method = values[2]
        values[1] = category
        values[2] = payment_method
        checklist_row = dict(zip(CHECKLIST_COLUMNS, values))
        checklist_row["coverage"] = "model_contract"
        rows.append(checklist_row)
    return rows


def build_fixture(performance_xlsx: Path, scenario_xlsx: Path) -> dict[str, Any]:
    block_rows, scenario_cases = load_scenario_cases(scenario_xlsx)
    checklist_rows = load_checklist_rows(performance_xlsx)
    block_counts = Counter(case["block_id"] for case in scenario_cases)
    expected_basket_counts = Counter(case.get("expected_basket") for case in scenario_cases)
    coverage_counts = Counter(case["coverage"] for case in scenario_cases)

    return {
        "metadata": {
            "schema_version": 1,
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "source_files": {
                "performance": performance_xlsx.name,
                "scenario_breakdown": scenario_xlsx.name,
            },
            "frame_stride": 1,
            "latency_budget_ms": 20_000,
            "expanded_case_count": len(scenario_cases),
            "checklist_row_count": len(checklist_rows),
            "product_catalog": {
                "A": {"product_id": 101, "name": "Scenario Product A", "weight": 101.0},
                "B": {"product_id": 102, "name": "Scenario Product B", "weight": 223.0},
                "C": {"product_id": 103, "name": "Scenario Product C", "weight": 359.0},
            },
            "coverage_counts": dict(sorted(coverage_counts.items())),
            "block_counts": dict(sorted(block_counts.items())),
            "expected_basket_counts": dict(sorted(expected_basket_counts.items())),
        },
        "blocks": block_rows,
        "expanded_cases": scenario_cases,
        "checklist_rows": checklist_rows,
    }


def write_report(fixture: dict[str, Any], path: Path) -> None:
    metadata = fixture["metadata"]
    block_counts = metadata["block_counts"]
    lines = [
        "# Scenario Fixture Report",
        "",
        f"Generated: {metadata['generated_at']}",
        "",
        "## Counts",
        "",
        f"- Expanded cases: {metadata['expanded_case_count']}",
        f"- Checklist rows: {metadata['checklist_row_count']}",
        f"- Frame stride contract: {metadata['frame_stride']}",
        f"- Latency budget: {metadata['latency_budget_ms']} ms",
        "",
        "## Block Counts",
        "",
    ]
    lines.extend(f"- {block_id}: {count}" for block_id, count in block_counts.items())
    lines.extend(
        [
            "",
            "## Coverage",
            "",
            "- Scenario rows are contract-tested against model-service basket judgment.",
            "- Registration, payment, card, and service-only checks remain external contracts.",
            "- Local PC tests do not prove Jetson TensorRT runtime readiness.",
            "",
        ]
    )
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines), encoding="utf-8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--performance-xlsx", required=True, type=Path)
    parser.add_argument("--scenario-xlsx", required=True, type=Path)
    parser.add_argument("--output", default=DEFAULT_OUTPUT, type=Path)
    parser.add_argument("--report", default=DEFAULT_REPORT, type=Path)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    fixture = build_fixture(args.performance_xlsx, args.scenario_xlsx)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(
        json.dumps(fixture, ensure_ascii=True, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )
    write_report(fixture, args.report)
    print(
        "wrote "
        f"{args.output} ({fixture['metadata']['expanded_case_count']} cases, "
        f"{fixture['metadata']['checklist_row_count']} checklist rows)"
    )


if __name__ == "__main__":
    main()
